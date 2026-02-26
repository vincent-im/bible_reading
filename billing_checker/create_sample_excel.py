"""샘플 Excel 파일 생성 스크립트"""
import os
import sys

# 현재 파일 위치를 기준으로 경로 설정
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl")
    sys.exit(1)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "청구내역"

# 헤더
headers = ["고객사명", "항목", "금액", "담당자", "담당자이메일", "고객사이메일", "청구월"]
header_fill = PatternFill("solid", fgColor="1A73E8")
header_font = Font(bold=True, color="FFFFFF")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# 데이터 (이메일과의 일치/불일치 케이스 포함)
data = [
    # (고객사명, 항목, 금액, 담당자, 담당자이메일, 고객사이메일, 청구월)
    # ABC상사: 이메일(2,000,000) = Excel(2,000,000) → 일치
    ("ABC상사 (주)", "클라우드 서비스 통합", 2_000_000, "홍길동", "manager1@mycompany.com", "billing@abc.co.kr", "2024-01"),
    # XYZ테크놀로지: 이메일(3,500,000) = Excel(3,500,000) → 일치
    ("XYZ테크놀로지", "SaaS 라이선스 + API", 3_500_000, "이영희", "manager2@mycompany.com", "finance@xyztech.com", "2024-01"),
    # 글로벌무역: 이메일(1,250,000) ≠ Excel(1,500,000) → 불일치
    ("글로벌무역(주)", "ERP 유지보수 패키지", 1_500_000, "박민수", "manager3@mycompany.com", "account@global-trade.com", "2024-01"),
    # 미래산업: Excel에만 있음 (이메일 누락 케이스)
    ("미래산업", "컨설팅 서비스", 800_000, "최지수", "manager4@mycompany.com", "admin@miraeind.com", "2024-01"),
]

fills = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="F8F9FA")]
for row_idx, row_data in enumerate(data, 2):
    fill = fills[(row_idx) % 2]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        if col_idx == 3:  # 금액 컬럼
            cell.number_format = '#,##0'

# 열 너비 조정
widths = [20, 25, 15, 10, 28, 28, 12]
from openpyxl.utils import get_column_letter
for col_idx, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

os.makedirs("sample_data", exist_ok=True)
output_path = "sample_data/total_billing.xlsx"
wb.save(output_path)
print(f"샘플 Excel 파일 생성 완료: {output_path}")
print(f"  - {len(data)}개 고객사 청구 데이터")
