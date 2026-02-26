"""
리포트 생성 모듈

비교 결과를 Excel 및 HTML 형식으로 저장합니다.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import List

from .models import ComparisonResult, ComparisonStatus

logger = logging.getLogger(__name__)

# 상태별 스타일 매핑
STATUS_STYLE = {
    ComparisonStatus.MATCH:           {"label": "일치",        "color": "#28a745", "bg": "#d4edda"},
    ComparisonStatus.MISMATCH:        {"label": "불일치",      "color": "#dc3545", "bg": "#f8d7da"},
    ComparisonStatus.EMAIL_NOT_FOUND: {"label": "이메일 없음", "color": "#fd7e14", "bg": "#fff3cd"},
    ComparisonStatus.EXCEL_NOT_FOUND: {"label": "Excel 없음",  "color": "#fd7e14", "bg": "#fff3cd"},
    ComparisonStatus.PARSE_ERROR:     {"label": "파싱 오류",   "color": "#6c757d", "bg": "#e2e3e5"},
}


class BillingReporter:
    """비교 결과를 Excel/HTML 리포트로 저장"""

    def __init__(self, cfg: dict):
        self.output_dir = Path(cfg.get("output_dir", "reports"))
        self.fmt = cfg.get("format", "both")
        self.prefix = cfg.get("filename_prefix", "billing_comparison")
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, results: List[ComparisonResult]) -> dict:
        """리포트 생성 후 저장된 파일 경로 반환"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        saved = {}

        if self.fmt in ("excel", "both"):
            path = self.output_dir / f"{self.prefix}_{ts}.xlsx"
            self._save_excel(results, path)
            saved["excel"] = str(path)
            logger.info("Excel 리포트 저장: %s", path)

        if self.fmt in ("html", "both"):
            path = self.output_dir / f"{self.prefix}_{ts}.html"
            self._save_html(results, path)
            saved["html"] = str(path)
            logger.info("HTML 리포트 저장: %s", path)

        return saved

    # ── Excel 리포트 ──────────────────────────────────────────────────────

    def _save_excel(self, results: List[ComparisonResult], path: Path):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl이 설치되어 있지 않습니다: pip install openpyxl")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "비교결과"

        # 헤더
        headers = [
            "고객사명", "청구월", "상태",
            "이메일 금액", "Excel 금액", "차이",
            "담당자", "담당자이메일", "고객사이메일",
            "이메일 제목", "조치사항", "처리일시",
        ]
        header_fill = PatternFill("solid", fgColor="1A73E8")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        ws.row_dimensions[1].height = 20

        # 데이터 행
        for row_idx, result in enumerate(results, 2):
            style = STATUS_STYLE.get(result.status, STATUS_STYLE[ComparisonStatus.PARSE_ERROR])
            row_fill = PatternFill("solid", fgColor=style["bg"].lstrip("#"))

            values = [
                result.client_name,
                result.billing_month,
                style["label"],
                result.email_amount,
                result.excel_amount,
                result.difference,
                result.manager_name,
                result.manager_email,
                result.client_email,
                result.email_subject,
                result.action_taken,
                result.processed_at.strftime("%Y-%m-%d %H:%M:%S"),
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                # 금액 컬럼 숫자 포맷
                if col_idx in (4, 5, 6) and isinstance(value, (int, float)):
                    cell.number_format = '#,##0'

        # 열 너비 자동 조정
        col_widths = [20, 12, 12, 16, 16, 14, 12, 28, 28, 40, 30, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # 요약 시트
        ws_summary = wb.create_sheet("요약")
        counts = {}
        for r in results:
            counts[r.status] = counts.get(r.status, 0) + 1

        ws_summary["A1"] = "청구 금액 비교 요약"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A3"] = "상태"
        ws_summary["B3"] = "건수"
        ws_summary["A3"].font = Font(bold=True)
        ws_summary["B3"].font = Font(bold=True)

        for i, (status, count) in enumerate(counts.items(), 4):
            style = STATUS_STYLE.get(status, {})
            ws_summary.cell(row=i, column=1, value=style.get("label", status.value))
            ws_summary.cell(row=i, column=2, value=count)

        ws_summary.cell(row=len(counts) + 5, column=1, value="합계")
        ws_summary.cell(row=len(counts) + 5, column=2, value=len(results))
        ws_summary.cell(row=len(counts) + 5, column=1).font = Font(bold=True)

        wb.save(path)

    # ── HTML 리포트 ───────────────────────────────────────────────────────

    def _save_html(self, results: List[ComparisonResult], path: Path):
        now = datetime.now().strftime("%Y년 %m월 %d일 %H:%M:%S")
        counts = {}
        for r in results:
            counts[r.status] = counts.get(r.status, 0) + 1

        def badge(status):
            s = STATUS_STYLE.get(status, {"label": "?", "color": "#333", "bg": "#eee"})
            return (
                f'<span style="background:{s["color"]};color:white;padding:3px 10px;'
                f'border-radius:10px;font-size:12px;">{s["label"]}</span>'
            )

        def fmt_amt(v):
            if v is None:
                return '<span style="color:#aaa">-</span>'
            return f"{v:,.0f}원"

        def fmt_diff(v):
            if v is None:
                return '<span style="color:#aaa">-</span>'
            color = "#dc3545" if v != 0 else "#28a745"
            sign = "+" if v > 0 else ""
            return f'<span style="color:{color};font-weight:bold">{sign}{v:,.0f}원</span>'

        rows_html = ""
        for r in results:
            rows_html += f"""<tr>
  <td>{r.client_name}</td>
  <td style="text-align:center">{r.billing_month or "-"}</td>
  <td style="text-align:center">{badge(r.status)}</td>
  <td style="text-align:right">{fmt_amt(r.email_amount)}</td>
  <td style="text-align:right">{fmt_amt(r.excel_amount)}</td>
  <td style="text-align:right">{fmt_diff(r.difference)}</td>
  <td>{r.manager_name or "-"}</td>
  <td>{r.action_taken or "-"}</td>
</tr>
"""

        summary_html = ""
        for status, cnt in counts.items():
            s = STATUS_STYLE.get(status, {"label": "?", "color": "#333", "bg": "#eee"})
            summary_html += (
                f'<span style="margin:4px;display:inline-block;padding:6px 16px;'
                f'border-radius:20px;background:{s["bg"]};border:1px solid {s["color"]};'
                f'color:{s["color"]};font-weight:bold">{s["label"]}: {cnt}건</span>'
            )

        html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="utf-8">
<title>청구 금액 비교 리포트</title>
<style>
  * {{ box-sizing: border-box; }}
  body {{ font-family: '맑은 고딕', Arial, sans-serif; color: #333; margin: 0; background: #f0f2f5; }}
  .container {{ max-width: 1200px; margin: 30px auto; background: white;
               border-radius: 12px; box-shadow: 0 2px 12px rgba(0,0,0,0.1); overflow: hidden; }}
  .header {{ background: linear-gradient(135deg, #1a73e8, #0d47a1); color: white; padding: 30px; }}
  .header h1 {{ margin: 0 0 6px; font-size: 24px; }}
  .header p {{ margin: 0; opacity: 0.85; }}
  .summary {{ padding: 20px 30px; background: #f8f9fa; border-bottom: 1px solid #dee2e6; }}
  .summary h3 {{ margin: 0 0 12px; font-size: 14px; color: #666; text-transform: uppercase; letter-spacing: 1px; }}
  .table-wrap {{ padding: 20px 30px; overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
  thead th {{ background: #1a73e8; color: white; padding: 12px 10px; text-align: left;
              font-weight: 600; white-space: nowrap; }}
  tbody tr:hover {{ background: #f0f4ff; }}
  tbody td {{ padding: 10px; border-bottom: 1px solid #dee2e6; vertical-align: middle; }}
  .footer {{ padding: 16px 30px; font-size: 12px; color: #999; border-top: 1px solid #dee2e6; }}
  @media print {{ body {{ background: white; }} .container {{ box-shadow: none; }} }}
</style></head>
<body><div class="container">
  <div class="header">
    <h1>청구 금액 비교 리포트</h1>
    <p>생성일시: {now} | 총 {len(results)}건</p>
  </div>
  <div class="summary">
    <h3>요약</h3>
    {summary_html}
  </div>
  <div class="table-wrap">
    <table>
      <thead><tr>
        <th>고객사명</th><th>청구월</th><th>상태</th>
        <th>이메일 금액</th><th>Excel 금액</th><th>차이</th>
        <th>담당자</th><th>조치사항</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
  </div>
  <div class="footer">
    본 리포트는 청구 금액 비교 자동화 시스템에서 자동 생성되었습니다.
  </div>
</div></body></html>"""

        path.write_text(html, encoding="utf-8")
